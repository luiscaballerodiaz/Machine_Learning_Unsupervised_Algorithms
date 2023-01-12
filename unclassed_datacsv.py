import pandas as pd
import numpy as np
import math
import time
import sys
from openpyxl import Workbook
import matplotlib.pyplot as plt
from sklearn.neighbors import KNeighborsClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import train_test_split
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler


class DataUnclassedCSV:
    """Class to operate with a dataset in CSV format"""

    def __init__(self, name):
        self.name = name
        self.fig_width = 20
        self.fig_height = 10
        self.bar_width = 0.25
        self.percentile = 0.02
        self.method = []
        self.parameters = []
        self.out_data = np.zeros([0, 8])
        self.pca = None
        self.X_train = None
        self.X_train_scaled = None
        self.X_train_scaled_pca = None
        self.X_test_scaled_pca = None
        self.X_train_scaled_pca_output0 = None
        self.X_train_scaled_pca_output1 = None
        self.X_test = None
        self.X_test_scaled = None
        self.y_train = None
        self.y_test = None
        self.y_train_output0 = None
        self.y_train_output1 = None
        self.y_test_output0 = None
        self.y_test_output1 = None

    def read_csv(self):
        """"Read csv and convert it to dataframe"""
        dataset = pd.read_csv(self.name)
        print("Full source data from CSV type: {} and shape: {}".format(type(dataset), dataset.shape))
        return dataset

    def unclassed_boxplot(self, dataset, plot_name, max_features_row):
        """Plot histogram based on input dataset"""
        dfcopy = dataset.copy()
        max_vector = np.zeros([dataset.shape[1]])
        for i in range(dataset.shape[1]):
            max_vector[i] = dataset.iloc[:, i].max()
        columns = []
        for i in range(dataset.shape[1]):
            index_max = np.argmax(max_vector)
            columns.append(dataset.columns.values[index_max])
            max_vector[index_max] = 0
        dfcopy = dfcopy.reindex(columns=columns)
        dfcopy = dfcopy.replace(np.nan, 0)
        fig, axes = plt.subplots(math.ceil(dataset.shape[1] / max_features_row), 1,
                                 figsize=(self.fig_width, self.fig_height))
        ax = axes.ravel()
        for i in range(len(ax)):
            ax[i].boxplot(dfcopy.iloc[:, (i * max_features_row):min(((i + 1) * max_features_row), dataset.shape[1])])
            ax[i].grid(visible=True)
            ax[i].tick_params(axis='both', labelsize=8)
            if ((i + 1) * max_features_row) > dataset.shape[1]:
                xrange = range(1, dataset.shape[1] - (i * max_features_row) + 1)
            else:
                xrange = range(1, max_features_row + 1)
            ax[i].set_xticks(xrange,
                             dfcopy.keys()[(i * max_features_row):min(((i + 1) * max_features_row), dataset.shape[1])],
                             rotation=10, ha='center')
            ax[i].set_ylabel('Feature magnitude', fontsize=8)
        ax[0].set_title(plot_name, fontsize=24, fontweight='bold')
        plt.savefig(plot_name + '.png', bbox_inches='tight')
        plt.clf()

    def unclassed_histogram(self, dataset, plot_name, ncolumns):
        """Plot histogram based on input dataset"""
        fig, axes = plt.subplots(math.ceil(dataset.shape[1] / ncolumns), ncolumns,
                                 figsize=(self.fig_width, self.fig_height))
        spare_axes = ncolumns - dataset.shape[1] % ncolumns
        if spare_axes == ncolumns:
            spare_axes = 0
        for axis in range(ncolumns - 1,  ncolumns - 1 - spare_axes, -1):
            fig.delaxes(axes[math.ceil(dataset.shape[1] / ncolumns) - 1, axis])
        ax = axes.ravel()
        for i in range(dataset.shape[1]):
            ax[i].hist(dataset.iloc[:, i], histtype='stepfilled', bins=25, alpha=0.25, color="#0000FF", lw=0)
            ax[i].set_title(dataset.keys()[i], fontsize=10, y=1.0, pad=-14, fontweight='bold')
            ax[i].grid(visible=True)
            ax[i].tick_params(axis='both', labelsize=8)
            ax[i].set_ylabel('Frequency', fontsize=8)
            ax[i].set_xlabel('Feature magnitude', fontsize=8)
        plt.savefig(plot_name + '.png', bbox_inches='tight')
        plt.clf()

    def data_scrubbing(self, dataset, columns_to_remove, concept1, concept2, encodings, class_column_name):
        """Scrub data from input dataset by removing the introduced columns, duplicates, empty and wrong values,
        and apply one hot encoding for categorical features"""
        # Remove non-meaningful columns
        dataset.drop(columns_to_remove, axis=1, inplace=True)
        print("Scrubber data after eliminating non-meaningful columns type: {} and shape: {}".format(type(dataset),
                                                                                                     dataset.shape))
        # One hot encoding
        output_backup = dataset[class_column_name]
        dataset.drop(class_column_name, axis=1, inplace=True)
        for encoding in encodings:
            dataset[encoding] = dataset[encoding].astype(str)
        dataset = pd.get_dummies(dataset, columns=encodings)
        dataset = pd.concat([dataset, output_backup], axis=1)
        print("Scrubber data after one hot encoding type: {} and shape: {}".format(type(dataset), dataset.shape))
        # Remove duplicates
        dataset.drop_duplicates(keep='first', inplace=True)
        print("Scrubber data after eliminating duplicates type: {} and shape: {}".format(type(dataset), dataset.shape))
        # Remove outliers
        df_qmin = dataset.quantile(self.percentile)
        df_qmax = dataset.quantile(1 - self.percentile)
        for i in range(len(dataset.keys())):
            if min(dataset.iloc[:, i]) == 0 and max(dataset.iloc[:, i] <= 1):
                continue
            else:
                dataset = dataset.loc[dataset[dataset.keys()[i]] >= df_qmin[i], :]
                dataset = dataset.loc[dataset[dataset.keys()[i]] <= df_qmax[i], :]
        print("Scrubber data after eliminating outliers type: {} and shape: {}".format(type(dataset), dataset.shape))
        # Remove empty rows
        dataset.replace('', np.nan, inplace=True)
        dataset.dropna(axis=0, how='any', inplace=True)
        dataset.reset_index(drop=True, inplace=True)
        print("Scrubber data after eliminating empty datasets type: {} and shape: {}".format(type(dataset),
                                                                                             dataset.shape))
        # Remove wrong rows if concept1 is higher than concept2
        index_to_drop = []
        for i in range(dataset.shape[0]):
            if dataset.iloc[i, dataset.columns.get_loc(concept1)] > dataset.iloc[i, dataset.columns.get_loc(concept2)]:
                index_to_drop.append(i)
        dataset.drop(index_to_drop, inplace=True)
        dataset.reset_index(drop=True, inplace=True)
        print("Scrubber data after eliminating non-consistent datasets type: {} and shape: {}".format(type(dataset),
                                                                                                      dataset.shape))
        return dataset

    def train_test_split(self, feature_data, class_data, test_size):
        """Split data into training and test datasets and plot the class distribution in each set"""
        self.X_train, self.X_test, self.y_train, self.y_test = train_test_split(feature_data, class_data,
                                                                                test_size=test_size, shuffle=True,
                                                                                stratify=class_data, random_state=1)
        self.y_train = np.ravel(self.y_train)
        self.y_test = np.ravel(self.y_test)
        print("X_train type: {} and shape: {}".format(type(self.X_train), self.X_train.shape))
        print("X_test type: {} and shape: {}".format(type(self.X_test), self.X_test.shape))
        print("y_train type: {} and shape: {}".format(type(self.y_train), self.y_train.shape))
        print("y_test type: {} and shape: {} \n".format(type(self.y_test), self.y_test.shape))
        self.plot_output_class_distribution()

    def plot_output_class_distribution(self):
        """Plot the class distribution in the training and test dataset"""
        self.y_train_output1 = self.y_train[self.y_train == 1]
        self.y_train_output0 = self.y_train[self.y_train == 0]
        self.y_test_output1 = self.y_test[self.y_test == 1]
        self.y_test_output0 = self.y_test[self.y_test == 0]
        print("y_train_output1 type: {} and shape: {}".format(type(self.y_train_output1), self.y_train_output1.shape))
        print("y_test_output1 type: {} and shape: {}".format(type(self.y_test_output1), self.y_test_output1.shape))
        print("y_train_output0 type: {} and shape: {}".format(type(self.y_train_output0), self.y_train_output0.shape))
        print("y_test_output0 type: {} and shape: {} \n".format(type(self.y_test_output0), self.y_test_output0.shape))

        plt.subplots(figsize=(self.fig_width, self.fig_height))
        plt.bar([1, 2], [self.y_train_output0.shape[0], self.y_test_output0.shape[0]],
                color='r', width=self.bar_width, edgecolor='black', label='class=0')
        plt.bar([1 + self.bar_width, 2 + self.bar_width], [self.y_train_output1.shape[0], self.y_test_output1.shape[0]],
                color='b', width=self.bar_width, edgecolor='black', label='class=1')
        plt.xticks([1 + self.bar_width / 2, 2 + self.bar_width / 2],
                   ['Train data', 'Test data'], ha='center')
        plt.text(1 - self.bar_width / 4, self.y_train_output0.shape[0] + 100,
                 str(self.y_train_output0.shape[0]), fontsize=20)
        plt.text(1 + 3 * self.bar_width / 4, self.y_train_output1.shape[0] + 100,
                 str(self.y_train_output1.shape[0]), fontsize=20)
        plt.text(2 - self.bar_width / 4, self.y_test_output0.shape[0] + 100,
                 str(self.y_test_output0.shape[0]), fontsize=20)
        plt.text(2 + 3 * self.bar_width / 4, self.y_test_output1.shape[0] + 100,
                 str(self.y_test_output1.shape[0]), fontsize=20)
        plt.title('Output class distribution between train and test datasets', fontsize=24)
        plt.xlabel('Concepts', fontweight='bold', fontsize=14)
        plt.ylabel('Count train/test class cases', fontweight='bold', fontsize=14)
        plt.legend()
        plt.grid()
        plt.savefig('Count_class_cases.png', bbox_inches='tight')
        plt.clf()

    def data_scaling(self, algorithm):
        """Scaling data to normalization or standardization"""
        if algorithm.lower() == 'norm':
            scaler = MinMaxScaler()
        elif algorithm.lower() == 'standard':
            scaler = StandardScaler()
        else:
            print('Algorithm not correct')
            return None
        scaler.fit(self.X_train)
        self.X_train_scaled = scaler.transform(self.X_train)
        self.X_test_scaled = scaler.transform(self.X_test)
        print("X_train_scaled type: {} and shape: {}".format(type(self.X_train_scaled), self.X_train_scaled.shape))
        print("X_test_scaled type: {} and shape: {} \n".format(type(self.X_test_scaled), self.X_test_scaled.shape))

    def apply_pca(self, ncomps):
        """Apply PCA algorithm in the scaled trained data and plot meaningful graphs"""
        self.pca = PCA(n_components=ncomps)
        self.pca.fit(self.X_train_scaled)
        self.X_train_scaled_pca = self.pca.transform(self.X_train_scaled)
        self.X_test_scaled_pca = self.pca.transform(self.X_test_scaled)
        print("X_train_scaled PCA type: {} and shape: {}".format(type(self.X_train_scaled_pca),
                                                                 self.X_train_scaled_pca.shape))
        self.X_train_scaled_pca_output1 = self.X_train_scaled_pca[self.y_train == 1, :]
        self.X_train_scaled_pca_output0 = self.X_train_scaled_pca[self.y_train == 0, :]
        print("X_train_scaled PCA output = 1 type: {} and shape: {}"
              .format(type(self.X_train_scaled_pca_output1), self.X_train_scaled_pca_output1.shape))
        print("X_train_scaled PCA output = 0 type: {} and shape: {}"
              .format(type(self.X_train_scaled_pca_output0), self.X_train_scaled_pca_output0.shape))
        print("PCA component shape: {} \n".format(self.pca.components_.shape))
        self.plot_pca_breakdown()
        self.plot_pca_scree()
        if ncomps > 2:
            self.plot_first_second_pca()

    def plot_pca_breakdown(self):
        """Plot the PCA breakdown per each feature"""
        _, ax = plt.subplots(figsize=(self.fig_width, self.fig_height))
        plt.pcolormesh(self.pca.components_, cmap=plt.cm.cool)
        plt.colorbar()
        pca_yrange = [x + 0.5 for x in range(self.pca.components_.shape[0])]
        pca_xrange = [x + 0.5 for x in range(self.pca.components_.shape[1])]
        try:
            plt.xticks(pca_xrange, self.X_train.keys(), rotation=60, ha='center')
        except (Exception,):
            str_xpca = []
            for i in range(self.pca.components_.shape[1]):
                str_xpca.append('Feature ' + str(i + 1))
            plt.xticks(pca_xrange, str_xpca, rotation=60, ha='center')
        ax.xaxis.tick_top()
        str_ypca = []
        for i in range(self.pca.components_.shape[0]):
            str_ypca.append('Component ' + str(i + 1))
        plt.yticks(pca_yrange, str_ypca)
        plt.xlabel("Feature", weight='bold', fontsize=14)
        plt.ylabel("Principal components", weight='bold', fontsize=14)
        plt.savefig('PCA_scaled_breakdown.png', bbox_inches='tight')
        plt.clf()

    def plot_pca_scree(self):
        """Plot the scree plot of the PCA to understand the covered variance"""
        fig, ax1 = plt.subplots(figsize=(self.fig_width, self.fig_height))
        ax2 = ax1.twinx()
        label1 = ax1.plot(range(1, len(self.pca.components_) + 1), self.pca.explained_variance_ratio_,
                          'ro-', linewidth=2, label='Individual PCA variance')
        label2 = ax2.plot(range(1, len(self.pca.components_) + 1), np.cumsum(self.pca.explained_variance_ratio_),
                          'b^-', linewidth=2, label='Cumulative PCA variance')
        plt.title('Scree Plot', fontsize=20, fontweight='bold')
        ax1.set_xlabel('Principal Components', fontsize=14)
        ax1.set_ylabel('Proportion of Variance Explained', fontsize=14, color='r')
        ax2.set_ylabel('Cumulative Proportion of Variance Explained', fontsize=14, color='b')
        la = label1 + label2
        lb = [la[0].get_label(), la[1].get_label()]
        ax1.legend(la, lb, loc='upper center')
        ax1.grid(visible=True)
        ax2.grid(visible=True)
        plt.savefig('PCA_scaled_ScreePlot.png', bbox_inches='tight')
        plt.clf()

    def plot_first_second_pca(self):
        """Plot first vs second PCA component"""
        plt.subplots(figsize=(self.fig_width, self.fig_height))
        plt.scatter(self.X_train_scaled_pca_output1[:, 0], self.X_train_scaled_pca_output1[:, 1],
                    s=10, marker='^', c='red', label='output=1')
        plt.scatter(self.X_train_scaled_pca_output0[:, 0], self.X_train_scaled_pca_output0[:, 1],
                    s=10, marker='o', c='blue', label='output=0')
        plt.title('Cardiovascular disease modelling', fontsize=20, fontweight='bold')
        plt.xlabel('First PCA', fontsize=14)
        plt.ylabel('Second PCA', fontsize=14)
        plt.legend()
        plt.grid()
        plt.savefig('PCA_scaled_First_Second.png', bbox_inches='tight')
        plt.clf()

    def apply_algorithm(self, algorithm, params):
        """Apply the machine learning algorithm to the train and test datasets"""
        time0 = time.time()
        if algorithm.lower() == 'knn':
            model = KNeighborsClassifier()
        elif algorithm.lower() == 'logreg':
            model = LogisticRegression(random_state=0)
        elif algorithm.lower() == 'linearsvc':
            model = LinearSVC(random_state=0)
        elif algorithm.lower() == 'naivebayes':
            model = GaussianNB()
        elif algorithm.lower() == 'tree':
            model = DecisionTreeClassifier(random_state=0)
        elif algorithm.lower() == 'forest':
            model = RandomForestClassifier(random_state=0)
        elif algorithm.lower() == 'gradient':
            model = GradientBoostingClassifier(random_state=0)
        elif algorithm.lower() == 'svm':
            model = SVC(random_state=0)
        elif algorithm.lower() == 'mlp':
            model = MLPClassifier(random_state=0)
        else:
            return None
        for key, value in params.items():
            setattr(model, key, value)
        print('SCORE WITH {} ALGORITHM AND PARAMS {}\n'.format(algorithm, params))
        model.fit(self.X_train, self.y_train)
        time1 = time.time()
        unscaled_model_time = round(time1 - time0, 4)
        unscaled_train_score = round(model.score(self.X_train, self.y_train), 4)
        unscaled_test_score = round(model.score(self.X_test, self.y_test), 4)
        print('Unscaled modeling time [seconds]: {}'.format(unscaled_model_time, 4))
        print('Unscaled TRAIN dataset: {}'.format(unscaled_train_score, 4))
        print('Unscaled TEST dataset: {}'.format(unscaled_test_score, 4))
        time2 = time.time()
        unscaled_predict_time = round(time2 - time1, 4)
        print('Unscaled predicting time [seconds]: {}\n'.format(unscaled_predict_time, 4))
        model.fit(self.X_train_scaled, self.y_train)
        time3 = time.time()
        scaled_model_time = round(time3 - time2, 4)
        scaled_train_score = round(model.score(self.X_train_scaled, self.y_train), 4)
        scaled_test_score = round(model.score(self.X_test_scaled, self.y_test), 4)
        print('Scaled modeling time [seconds]: {}'.format(scaled_model_time, 4))
        print('Scaling TRAIN dataset: {}'.format(scaled_train_score, 4))
        print('Scaling TEST dataset: {}'.format(scaled_test_score, 4))
        time4 = time.time()
        scaled_predict_time = round(time4 - time3, 4)
        print('Scaled predicting time [seconds]: {}\n\n'.format(scaled_predict_time, 4))
        results = np.array([[unscaled_model_time, unscaled_predict_time, unscaled_train_score, unscaled_test_score,
                             scaled_model_time, scaled_predict_time, scaled_train_score, scaled_test_score]])
        self.out_data = np.append(self.out_data, results, axis=0)
        self.method.append(algorithm)
        self.parameters.append(params)

    def write_results_excel_file(self, name):
        """Write the simulation results in an output excel file"""
        # Create excel file with the corresponding sheets
        sheets = []
        wb = Workbook()
        sheets.append(wb.active)
        sheets[0].title = 'SIMULATION RESULTS'
        # Define column width
        for column in range(1, 11):
            column_char = str(chr(64 + column))
            if column == 2:
                sheets[0].column_dimensions[column_char].width = 60
            else:
                sheets[0].column_dimensions[column_char].width = 20
        # Write headers
        header = ['Algorithm', 'Params', 'Unscaled Model Time', 'Unscaled Predict Time', 'Unscaled Train Score',
                  'Unscaled Test Score', 'Scaled Model Time', 'Scaled Predict Time', 'Scaled Train Score',
                  'Scaled Test Score']
        for i in range(len(header)):
            sheets[0].cell(1, i + 1).value = header[i]
        # Write algorithms
        for i in range(len(self.method)):
            sheets[0].cell(i + 2, 1).value = self.method[i]
        # Write parameters
        for i in range(len(self.parameters)):
            str_params = ''
            for key, value in self.parameters[i].items():
                str_params += ' ' + key + '=' + str(value)
            sheets[0].cell(i + 2, 2).value = str_params
        # Write data in excel sheet
        for i in range(self.out_data.shape[0]):
            for j in range(self.out_data.shape[1]):
                sheets[0].cell(i + 2, j + 3).value = self.out_data[i, j]
        try:
            wb.save(name)
        except PermissionError:
            sys.exit('ERROR: Excel file open. Please close it to be modified')
